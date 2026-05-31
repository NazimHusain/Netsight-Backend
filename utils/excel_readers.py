
import openpyxl
import ipaddress

def read_ips_from_excel(file_path):
    """
    Reads IPs from the 'Node IP' (2nd column) of Excel file.
    Validates IPv4, removes duplicates, ignores blanks.
    """

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    ips = set()  # avoid duplicates

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_ip = row[0]   # Column A: Node IP

        if not raw_ip:
            continue

        ip_str = str(raw_ip).strip()

        # Validate IP
        try:
            ipaddress.ip_address(ip_str)
            ips.add(ip_str)
        except ValueError:
            print(f"Invalid IP skipped: {ip_str}")
            continue

    wb.close()
    return list(ips)


def Ip_specific_command_reader(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    tasks = []
    sequence = 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_ip = row[0]
        raw_cmd = row[1]

        # Skip empty rows properly
        if raw_ip is None or raw_cmd is None:
            continue

        ip = str(raw_ip).strip()
        cmd = str(raw_cmd).strip()

        if not ip or not cmd:
            continue

        tasks.append({
            "ip": ip,
            "command": cmd,
            "sequence": sequence
        })
        sequence += 1

    wb.close()
    return tasks




def read_commands(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    commands = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) > 1 and row[1]:
            cmd = str(row[1]).strip()
            if cmd and cmd not in commands:
                commands.append(cmd)

    wb.close()
    return commands




def Ip_related_command_reader(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    tasks = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_ip = row[0]
        raw_cmd = row[1]

        if raw_ip is None or raw_cmd is None:
            continue

        ip = str(raw_ip).strip()
        cmd_block = str(raw_cmd).strip()

        if not ip or not cmd_block:
            continue

        commands = [c.strip() for c in cmd_block.split("\n") if c.strip()]

        if ip not in tasks:
            tasks[ip] = []

        tasks[ip].extend(commands)

    wb.close()
    return tasks
